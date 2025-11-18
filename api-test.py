# -*- coding: utf-8 -*-
from google import genai
from google.genai import types
import os
import logging

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Excel 파일 읽기
try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        import openpyxl

        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False


# 파일 읽기 함수 (여러 인코딩 시도)
def read_file_with_encoding(file_path: str) -> str:
    """파일을 읽어서 반환 (여러 인코딩 시도)"""
    encodings = ["utf-8", "utf-8-sig", "cp949", "euc-kr", "latin1"]

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, FileNotFoundError):
            continue

    raise FileNotFoundError(
        f"파일을 찾을 수 없거나 인코딩을 확인할 수 없습니다: {file_path}"
    )


# Excel 파일 읽기 함수
def read_excel_file(file_path: str) -> str:
    """Excel 파일을 읽어서 텍스트로 반환"""
    if HAS_PANDAS:
        try:
            excel_file = pd.ExcelFile(file_path)
            content_parts = []

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                content_parts.append(f"[시트: {sheet_name}]\n")
                content_parts.append(df.to_string(index=False))
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Excel 파일을 읽는 중 오류 발생: {e}")
    elif HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            content_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content_parts.append(f"[시트: {sheet_name}]\n")

                for row in ws.iter_rows(values_only=True):
                    row_str = "\t".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    content_parts.append(row_str)
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Excel 파일을 읽는 중 오류 발생: {e}")
    else:
        raise ImportError(
            "Excel 파일을 읽으려면 pandas 또는 openpyxl이 필요합니다. 'pip install pandas openpyxl'을 실행하세요."
        )


# 시스템 프롬프트 파일 읽기
def load_system_prompt(file_path: str = "prompts/evaluation_prompt.txt") -> str:
    """시스템 프롬프트 파일을 읽어서 반환"""
    content = read_file_with_encoding(file_path)
    # UTF-8로 다시 저장하여 다음번에는 UTF-8로 읽을 수 있도록 함
    try:
        with open(file_path, "w", encoding="utf-8") as f_out:
            f_out.write(content)
    except:
        pass  # 저장 실패해도 계속 진행
    return content


# api_key = os.getenv("GOOGLE_API_KEY")
api_key = "AIzaSyDlyiS2521pfi_pgJngVS0sUnnVT1v0n9A"

if not api_key:
    raise ValueError(
        "GOOGLE_API_KEY environment variable is not set. Please set it or provide the API key directly."
    )

# 시스템 프롬프트 로드
logger.info("시스템 프롬프트 로드 중...")
system_prompt = load_system_prompt()
logger.info("시스템 프롬프트 로드 완료")

# 다운로드 폴더에서 결과보고서 파일 읽기
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
report_file = None

# 읽을 파일명 (변경 가능)
target_file_keywords = ["9월", "스터디", "이용호"]  # 파일명에 포함되어야 할 키워드

# 다운로드 폴더에서 파일명으로 검색
if os.path.exists(downloads_path):
    for file in os.listdir(downloads_path):
        # 모든 키워드가 파일명에 포함되어 있는지 확인
        if all(keyword in file for keyword in target_file_keywords):
            report_file = os.path.join(downloads_path, file)
            logger.info(f"파일 발견: {file}")
            break

if not report_file or not os.path.exists(report_file):
    logger.error(f"파일을 찾을 수 없습니다: {downloads_path}")
    print(f"파일을 찾을 수 없습니다: {downloads_path}")
    print("다운로드 폴더에서 파일을 확인하거나 직접 경로를 지정해주세요.")
    report_content = "한국에 대해 알려줘"  # 기본값
else:
    try:
        # 파일 확장자에 따라 적절한 함수 사용
        file_ext = os.path.splitext(report_file)[1].lower()

        if file_ext in [".xlsx", ".xls"]:
            logger.info(f"Excel 파일 읽기: {report_file}")
            report_content = read_excel_file(report_file)
            print(f"Excel 파일을 읽었습니다: {report_file}")
        else:
            logger.info(f"텍스트 파일 읽기: {report_file}")
            report_content = read_file_with_encoding(report_file)
            print(f"파일을 읽었습니다: {report_file}")
    except Exception as e:
        logger.error(f"파일 읽기 실패: {e}")
        print(f"파일을 읽는 중 오류 발생: {e}")
        report_content = "한국에 대해 알려줘"  # 기본값

# Google AI API 호출
logger.info("Google AI API 호출 중...")
client = genai.Client(api_key=api_key)

try:
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        config=types.GenerateContentConfig(
            system_instruction=system_prompt,
        ),
        contents=report_content,
    )
    logger.info("API 호출 성공")
except Exception as e:
    logger.error(f"API 호출 실패: {e}")
    raise

print(response.text)
logger.info("완료")
