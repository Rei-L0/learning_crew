# file_utils.py
import os
import logging

# Excel 파일 읽기 라이브러리 임포트 시도
try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        import openpyxl

        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


def read_file_with_encoding(file_path: str) -> str:
    """파일을 읽어서 반환 (여러 인코딩 시도)"""
    encodings = ["utf-8", "utf-8-sig", "cp949", "euc-kr", "latin1"]

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, FileNotFoundError):
            continue

    logger.error(f"파일을 찾을 수 없거나 인코딩을 확인할 수 없습니다: {file_path}")
    raise FileNotFoundError(
        f"파일을 찾을 수 없거나 인코딩을 확인할 수 없습니다: {file_path}"
    )


def read_excel_file(file_path: str) -> str:
    """Excel 파일을 읽어서 텍스트로 반환"""
    if HAS_PANDAS:
        try:
            excel_file = pd.ExcelFile(file_path)
            content_parts = []

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                content_parts.append(f"[시트: {sheet_name}]\n")
                content_parts.append(df.to_string(index=False))
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Pandas로 Excel 파일을 읽는 중 오류 발생: {e}")

    elif HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            content_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content_parts.append(f"[시트: {sheet_name}]\n")

                for row in ws.iter_rows(values_only=True):
                    row_str = "\t".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    content_parts.append(row_str)
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Openpyxl로 Excel 파일을 읽는 중 오류 발생: {e}")
    else:
        raise ImportError(
            "Excel 파일을 읽으려면 pandas 또는 openpyxl이 필요합니다. 'pip install pandas openpyxl'을 실행하세요."
        )


def load_system_prompt(file_path: str) -> str:
    """시스템 프롬프트 파일을 읽고, UTF-8로 다시 저장 (원본 기능 유지)"""
    content = read_file_with_encoding(file_path)  # FileNotFoundError 여기서 발생 가능
    try:
        # UTF-8로 다시 저장하여 다음번에는 UTF-8로 읽을 수 있도록 함
        with open(file_path, "w", encoding="utf-8") as f_out:
            f_out.write(content)
    except Exception as e:
        logger.warning(f"시스템 프롬프트를 UTF-8로 다시 저장하는 데 실패했습니다: {e}")
    return content


def find_file_by_keywords(search_path: str, keywords: list[str]) -> str | None:
    """지정된 경로에서 키워드가 모두 포함된 파일명을 찾아 전체 경로를 반환합니다."""
    if not os.path.exists(search_path):
        logger.warning(f"검색 경로를 찾을 수 없습니다: {search_path}")
        return None

    for file in os.listdir(search_path):
        if all(keyword in file for keyword in keywords):
            logger.info(f"파일 발견: {file}")
            return os.path.join(search_path, file)

    logger.warning(
        f"키워드 {keywords}에 해당하는 파일을 {search_path}에서 찾지 못했습니다."
    )
    return None


def get_file_content(file_path: str) -> str:
    """파일 확장자에 따라 적절한 읽기 함수를 호출하여 내용을 반환합니다."""
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext in [".xlsx", ".xls"]:
        logger.info(f"Excel 파일 읽기 시도: {file_path}")
        return read_excel_file(file_path)
    else:
        logger.info(f"텍스트 파일 읽기 시도: {file_path}")
        return read_file_with_encoding(file_path)
